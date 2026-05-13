from fastapi.responses import StreamingResponse
from io import BytesIO
import pandas as pd

from fastapi import APIRouter, UploadFile, File, HTTPException, Request
from urllib.parse import quote
from fastapi.responses import FileResponse
from datetime import datetime
from pydantic import BaseModel
from typing import Optional, Union, Dict
import os
from src.logger.logger import get_logger
from src.services.excel_service import excel_service
from src.services.pdf_service import PapsPdfGenerator, StudentData
from src.database.db_handler import (
    insert_upload_file,
    insert_user_info_many,
    get_users_by_school,
    get_measurements_by_tags,
    get_user_by_search,
    get_measurements_by_user,
    insert_measurement,
    get_db
)


db = None

def get_database():
    global db
    if db is None:
        db = get_db()
    return db

logger = get_logger('API')
router = APIRouter()


# 헬퍼 함수: 측정값이 0이면 5등급 반환
def get_grade(value, grade):
    if value == 0 or value is None:
        return 5
    return grade if grade else 5


# 업로드 경로
UPLOAD_PATH = os.getenv("UPLOAD_PATH", "./uploads")

# 전역 변수
stored_email = None
stored_file_id = None
stored_origin_file_name = None


class EmailData(BaseModel):
    email: str
    file_id: str = None


class ExportRequest(BaseModel):
    message: str = None
    file_id: str = None
    origin_file_name: str = None


class MeasurementData(BaseModel):
    tag_number: str
    exercise_type: str
    value: Union[float, Dict]  # 숫자 또는 객체
    unit: Optional[str] = ""
    gender: Optional[str] = ""
    age: Optional[int] = 0
    grade: Optional[Union[int, Dict]] = None  # 프론트에서 계산한 등급
    timestamp: str


# PDF 검색 요청 모델
class PdfSearchRequest(BaseModel):
    measure_date: Optional[str] = None
    school_name: Optional[str] = None
    grade_year: Optional[str] = None
    class_number: Optional[str] = None
    student_number: Optional[str] = None
    student_name: Optional[str] = None


# ============================================
# 1. 엑셀 파일 업로드
# ============================================

@router.post("/api/upload/excel")
async def upload_excel(request: Request, file: UploadFile = File(...)):
    """학생 명단 엑셀 업로드 → DB 저장"""
    try:
        logger.info(f"파일 수신: {file.filename}")

        # 파일 검증
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="엑셀 파일만 업로드 가능합니다")

        # 파일 내용 읽기
        contents = await file.read()
        file_bytes = BytesIO(contents)

        # ✅ 1. 엑셀 검증
        validation_result = excel_service.excel_data_validate(file_bytes)
        file_bytes.seek(0)

        if not validation_result["valid"]:
            logger.warning(f"엑셀 검증 실패 : {validation_result}")
            raise HTTPException(
                status_code=400,
                detail={
                    "message": validation_result["message"],
                    "errors": validation_result.get("errors", [])
                }
            )

        # ✅ 2. user_info 데이터 추출 및 저장
        user_data = excel_service.extract_user_data(file_bytes)
        file_bytes.seek(0)

        if not user_data["success"]:
            raise HTTPException(status_code=400, detail=user_data["message"])

        # user_info DB 저장
        if user_data["users"]:
            insert_user_info_many(user_data["users"])
            logger.info(f"✅ user_info 저장 완료: {user_data['count']}명")

        # ✅ 3. upload_files 기록 생성 및 저장
        upload_data = excel_service.create_upload_record(file_bytes, file.filename)
        file_bytes.seek(0)

        if not upload_data["success"]:
            raise HTTPException(status_code=400, detail=upload_data["message"])

        file_id = insert_upload_file(upload_data["record"])
        logger.info(f"✅ upload_files 저장 완료, _id: {file_id}")

        # ✅ 4. 파일 저장
        saved_filename = upload_data["record"]["file_name"]
        file_path = os.path.join(UPLOAD_PATH, saved_filename)
        with open(file_path, "wb") as f:
            f.write(contents)
        logger.info(f"✅ 파일 저장: {file_path}")

        return {
            "success": True,
            "message": "학생 명단 업로드 완료",
            "file_name": saved_filename,
            "file_id": str(file_id),
            "user_count": user_data["count"],
            "school_code": upload_data["record"]["school_code"],
            "measure_date": validation_result["meta_data"]["측정일"].strftime("%Y-%m-%d")
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"업로드 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# 2. 데이터 export (세 개 시트)
# ============================================

@router.get("/api/download/excel/{school_code}/{file_name}")
async def download_excel(school_code: str, file_name: str):
    """업로드한 원본 엑셀에 측정 데이터 채워서 다운로드"""
    try:
        # 1. 원본 파일 경로
        file_path = os.path.join(UPLOAD_PATH, file_name)

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        # ============================================
        # 2. user_info에서 해당 학교 학생들의 tag_number 조회
        # ============================================
        users = get_users_by_school(school_code)
        if not users:
            raise HTTPException(status_code=404, detail="해당 학교의 학생 정보가 없습니다")

        tag_numbers = [str(user["tag_number"]) for user in users]
        logger.info(f"학교 {school_code}의 태그번호 목록: {tag_numbers}")

        # tag_number -> user 매핑 (이름 조회용)
        tag_to_user = {str(user["tag_number"]): user for user in users}

        # ============================================
        # 3. 엑셀에서 측정일 읽기
        # ============================================
        ws_value = wb["기본-1"]

        measure_cell = ws_value.cell(row=7, column=2)  # 7행 B열 = 측정일
        excel_measure_date = measure_cell.value

        # 측정일을 date 객체로 변환
        if isinstance(excel_measure_date, datetime):
            target_date = excel_measure_date.date()
        else:
            target_date = datetime.strptime(str(excel_measure_date), "%Y-%m-%d").date()

        logger.info(f"엑셀 측정일: {target_date}")

        # ============================================
        # 4. measurement_data에서 tag_number로 조회
        # ============================================
        measurements = get_measurements_by_tags(tag_numbers, target_date)

        # 태그번호별로 측정 데이터 정리 (값 + 등급)
        measurement_dict = {}
        for m in measurements:
            tag = str(m["tag_number"])
            if tag not in measurement_dict:
                measurement_dict[tag] = {"values": {}, "grades": {}}

            m_type = m["exercise_type"]  # 카멜케이스 그대로 사용
            measurement_dict[tag]["values"][m_type] = m.get("value")
            measurement_dict[tag]["grades"][m_type] = m.get("grade")

        logger.info(f"측정 데이터: {len(measurement_dict)}명")

        # ============================================
        # 5. 기본-1 시트 - 측정값 채우기 (카멜케이스)
        # ============================================
        column_mapping_value = {
            "50mRun": 4,  # D열
            "shuttleRun": 5,  # E열
            "sitAndReach": 8,  # H열
            "rollingUp": 9,  # I열
            "longJump": 10,  # J열
            "height": 11,  # K열
            "weight": 12,  # L열
        }

        # 번호 -> 태그번호 매핑 (나이스 양식에서 사용)
        number_to_tag = {}

        for row in range(10, ws_value.max_row + 1):
            number_cell = ws_value.cell(row=row, column=1)  # A열 = 번호
            tag_cell = ws_value.cell(row=row, column=2)  # B열 = 태그번호
            if tag_cell.value is None:
                continue

            tag = str(int(float(tag_cell.value)))  # 6.0 -> 6 변환
            number = str(number_cell.value).strip() if number_cell.value else None

            if number:
                number_to_tag[number] = tag

            if tag in measurement_dict:
                values = measurement_dict[tag]["values"]

                # gripStrength 처리 (leftGrip, rightGrip으로 분리)
                if "gripStrength" in values:
                    grip_value = values["gripStrength"]
                    if isinstance(grip_value, dict):
                        if grip_value.get("leftGrip"):
                            ws_value.cell(row=row, column=6).value = grip_value["leftGrip"]  # F열
                        if grip_value.get("rightGrip"):
                            ws_value.cell(row=row, column=7).value = grip_value["rightGrip"]  # G열

                # 나머지 종목 처리
                for m_type, col_num in column_mapping_value.items():
                    if m_type in values and values[m_type] is not None:
                        ws_value.cell(row=row, column=col_num).value = values[m_type]

        # ============================================
        # 6. 기본-2 시트 - 번호/태그번호/이름 복사 + 등급 채우기
        # ============================================
        if "기본-2" in wb.sheetnames:
            ws_grade = wb["기본-2"]

            column_mapping_grade = {
                "50mRun": 4,  # D열
                "shuttleRun": 5,  # E열
                "sitAndReach": 8,  # H열
                "rollingUp": 9,  # I열
                "longJump": 10,  # J열
                "bmi": 11,  # K열 (BMI 등급)
            }

            # 기본-1의 데이터 행 (10행부터)을 기본-2의 3행부터 복사
            grade_row = 3
            for value_row in range(10, ws_value.max_row + 1):
                tag_cell = ws_value.cell(row=value_row, column=2)
                if tag_cell.value is None:
                    continue

                # 번호, 태그번호, 이름 복사
                ws_grade.cell(row=grade_row, column=1).value = ws_value.cell(row=value_row, column=1).value  # 번호
                ws_grade.cell(row=grade_row, column=2).value = ws_value.cell(row=value_row, column=2).value  # 태그번호
                ws_grade.cell(row=grade_row, column=3).value = ws_value.cell(row=value_row, column=3).value  # 이름

                tag = str(int(float(tag_cell.value)))

                # 등급 채우기
                if tag in measurement_dict:
                    grades = measurement_dict[tag]["grades"]

                    # gripStrength 등급 처리
                    if "gripStrength" in grades:
                        grip_grade = grades["gripStrength"]
                        if isinstance(grip_grade, dict):
                            if grip_grade.get("leftGrip"):
                                ws_grade.cell(row=grade_row, column=6).value = grip_grade["leftGrip"]  # F열
                            if grip_grade.get("rightGrip"):
                                ws_grade.cell(row=grade_row, column=7).value = grip_grade["rightGrip"]  # G열

                    # 나머지 종목 등급
                    for m_type, col_num in column_mapping_grade.items():
                        if m_type in grades and grades[m_type] is not None:
                            ws_grade.cell(row=grade_row, column=col_num).value = grades[m_type]

                grade_row += 1

        # ============================================
        # 7. 나이스 양식 시트 - 이름 + 값 채우기
        # ============================================
        if "나이스 양식" in wb.sheetnames:
            ws_neis = wb["나이스 양식"]

            for row in range(3, ws_neis.max_row + 1):
                number_cell = ws_neis.cell(row=row, column=1)  # A열 = 번호
                if number_cell.value is None:
                    continue

                number = str(number_cell.value).strip()
                tag = number_to_tag.get(number)

                if not tag:
                    continue

                # 이름 복사 (user_info에서)
                if tag in tag_to_user:
                    ws_neis.cell(row=row, column=2).value = tag_to_user[tag]["name"]

                if tag not in measurement_dict:
                    continue

                values = measurement_dict[tag]["values"]

                # 왕복오래달리기 - C열 (3)
                if "shuttleRun" in values:
                    ws_neis.cell(row=row, column=3).value = values["shuttleRun"]

                # 앉아윗몸앞으로굽히기 - G열 (7)
                if "sitAndReach" in values:
                    ws_neis.cell(row=row, column=7).value = values["sitAndReach"]

                # 윗몸말아올리기 - N열 (14)
                if "rollingUp" in values:
                    ws_neis.cell(row=row, column=14).value = values["rollingUp"]

                # 악력 - O열 (15) - MAX(좌, 우)
                if "gripStrength" in values:
                    grip_value = values["gripStrength"]
                    if isinstance(grip_value, dict):
                        left_grip = grip_value.get("leftGrip", 0) or 0
                        right_grip = grip_value.get("rightGrip", 0) or 0
                        max_grip = grip_value.get("maxGrip") or max(left_grip, right_grip)
                        ws_neis.cell(row=row, column=15).value = max_grip

                # 50m달리기 - P열 (16)
                if "50mRun" in values:
                    ws_neis.cell(row=row, column=16).value = values["50mRun"]

                # 제자리멀리뛰기 - Q열 (17)
                if "longJump" in values:
                    ws_neis.cell(row=row, column=17).value = values["longJump"]

                # BMI - R열 (18) - 계산: 몸무게 / (키/100)²
                height = values.get("height")
                weight = values.get("weight")
                if height and weight and height > 0:
                    height_m = height / 100
                    bmi = round(weight / (height_m ** 2), 1)
                    ws_neis.cell(row=row, column=18).value = bmi

        # ============================================
        # 8. BytesIO로 저장 및 응답
        # ============================================
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        # 원본 파일명으로 다운로드
        original_filename = file_name.split('_', 1)[-1] if '_' in file_name else file_name

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{quote(original_filename)}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"다운로드 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# 3. 측정값 데이터 receive (프론트에서 받은 값 그대로 저장)
# ============================================
@router.post("/api/measure")
async def sensor_endpoint(data: MeasurementData):
    try:
        logger.info(f"📥 Receive Data: {data}")

        # dict로 변환 - 프론트에서 받은 값 그대로 저장
        data_dict = data.dict(exclude_none=True)

        logger.info(
            f"📊 저장 데이터: exercise_type={data_dict.get('exercise_type')}, value={data_dict.get('value')}, grade={data_dict.get('grade')}")

        # DB에 저장 (insert or update)
        result = insert_measurement(data_dict)

        if result:
            return {
                "status": "success",
                "message": "측정 데이터 저장 완료",
                "action": result.get("action"),
                "id": result.get("id")
            }
        else:
            raise HTTPException(status_code=500, detail="DB 저장 실패")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ 측정 데이터 수신 중 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# 4. PDF 다운로드 (개인 출력)
# ============================================
@router.post("/api/pdf")
async def print_individual_pdf(request: PdfSearchRequest):
    """개인 PDF 출력 - 태그번호로 조회"""
    try:
        logger.info(f"PDF 생성 요청: tag_number={request.tag_number}, school_code={request.school_code}, date={request.measure_date}")

        # 1. user_info에서 학생 검색 (school_code + tag_number)
        query = {"tag_number": str(request.tag_number)}
        if request.school_code:
            query["school_code"] = request.school_code

        user = db.user_info.find_one(query)

        if not user:
            raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다.")

        logger.info(f"학생 조회 완료: {user['name']}, tag_number: {user['tag_number']}")

        # 2. measurement_data에서 측정 데이터 검색
        measurement = get_measurements_by_user(
            tag_number=user["tag_number"],
            age=user["age"],
            measure_date=request.measure_date
        )

        if not measurement:
            raise HTTPException(status_code=404, detail="측정 데이터를 찾을 수 없습니다.")

        logger.info(f"측정 데이터 조회 완료: {measurement}")

        # 3. BMI 계산
        height = measurement.get("height", 0) or 0
        weight = measurement.get("weight", 0) or 0
        bmi = 0
        if height and weight and height > 0:
            height_m = height / 100
            bmi = round(weight / (height_m ** 2), 2)

        # 4. 악력 데이터 처리
        grip_data = measurement.get("gripStrength", {})
        if isinstance(grip_data, dict):
            grip_right = grip_data.get("rightGrip", 0) or 0
            grip_left = grip_data.get("leftGrip", 0) or 0
        else:
            grip_right = 0
            grip_left = 0

        # 5. 악력 등급 처리
        grip_grade = measurement.get("grade_gripStrength", {})
        if isinstance(grip_grade, dict):
            grade_grip_right = grip_grade.get("rightGrip")
            grade_grip_left = grip_grade.get("leftGrip")
        else:
            grade_grip_right = None
            grade_grip_left = None

        # 6. 종합 등급 계산
        all_grades = [
            get_grade(measurement.get("50mRun"), measurement.get("grade_50mRun")),
            get_grade(measurement.get("shuttleRun"), measurement.get("grade_shuttleRun")),
            get_grade(measurement.get("rollingUp"), measurement.get("grade_rollingUp")),
            get_grade(measurement.get("sitAndReach"), measurement.get("grade_sitAndReach")),
            get_grade(measurement.get("longJump"), measurement.get("grade_longJump")),
            get_grade(grip_right, grade_grip_right),
            get_grade(grip_left, grade_grip_left),
            get_grade(bmi, measurement.get("grade_bmi")),
        ]
        valid_grades = [g for g in all_grades if g and g > 0]
        grade_total = round(sum(valid_grades) / len(valid_grades)) if valid_grades else 5

        # 7. StudentData 생성
        student = StudentData(
            name=user["name"],
            age=user["age"],
            gender="남" if measurement.get("gender") == "M" else "여",
            school="",
            class_info=f"{user.get('grade_year', '')}학년 {user.get('class_number', '')}반 {user.get('student_number', '')}번",
            height=height,
            weight=weight,
            bmi=bmi,
            run_50m=measurement.get("50mRun", 0) or 0,
            shuttle_run=measurement.get("shuttleRun", 0) or 0,
            sit_ups=measurement.get("rollingUp", 0) or 0,
            flexibility=measurement.get("sitAndReach", 0) or 0,
            long_jump=measurement.get("longJump", 0) or 0,
            grip_right=grip_right,
            grip_left=grip_left,
            grade_50m=get_grade(measurement.get("50mRun"), measurement.get("grade_50mRun")),
            grade_shuttle=get_grade(measurement.get("shuttleRun"), measurement.get("grade_shuttleRun")),
            grade_situps=get_grade(measurement.get("rollingUp"), measurement.get("grade_rollingUp")),
            grade_flexibility=get_grade(measurement.get("sitAndReach"), measurement.get("grade_sitAndReach")),
            grade_long_jump=get_grade(measurement.get("longJump"), measurement.get("grade_longJump")),
            grade_grip_right=get_grade(grip_right, grade_grip_right),
            grade_grip_left=get_grade(grip_left, grade_grip_left),
            grade_bmi=get_grade(bmi, measurement.get("grade_bmi")),
            grade_total=grade_total,
        )

        # 8. PDF 생성
        os.makedirs("src/output", exist_ok=True)
        output_path = f"src/output/{user['name']}_PAPS결과지.pdf"

        generator = PapsPdfGenerator(
            template_path="src/templates/template.pdf",
            font_path="src/fonts"
        )
        generator.generate(student, output_path)

        logger.info(f"PDF 생성 완료: {output_path}")

        # 9. 파일 응답
        return FileResponse(
            path=output_path,
            filename=f"{user['name']}_PAPS결과지.pdf",
            media_type="application/pdf"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF 생성 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# 5. PDF 다운로드 (전체 출력)
# ============================================
@router.get("/api/pdf/all/{school_code}/{measure_date}")
async def print_all_pdf(school_code: str, measure_date: str):
    """전체 학생 PDF 출력 - ZIP으로 다운로드"""
    import zipfile
    from io import BytesIO

    try:
        logger.info(f"전체 PDF 생성 요청: school_code={school_code}, date={measure_date}")

        # 1. 해당 학교 학생 조회
        users = get_users_by_school(school_code)
        if not users:
            raise HTTPException(status_code=404, detail="해당 학교의 학생 정보가 없습니다")

        # 2. ZIP 파일 생성
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            pdf_count = 0

            for user in users:
                # 3. 각 학생의 측정 데이터 조회
                measurement = get_measurements_by_user(
                    tag_number=user["tag_number"],
                    age=user["age"],
                    measure_date=measure_date
                )

                if not measurement:
                    logger.warning(f"측정 데이터 없음: {user['name']}")
                    continue

                # 4. BMI 계산
                height = measurement.get("height", 0) or 0
                weight = measurement.get("weight", 0) or 0
                bmi = 0
                if height and weight and height > 0:
                    height_m = height / 100
                    bmi = round(weight / (height_m ** 2), 2)

                # 5. 악력 데이터 처리
                grip_data = measurement.get("gripStrength", {})
                if isinstance(grip_data, dict):
                    grip_right = grip_data.get("rightGrip", 0) or 0
                    grip_left = grip_data.get("leftGrip", 0) or 0
                else:
                    grip_right = 0
                    grip_left = 0

                # 6. 악력 등급 처리
                grip_grade = measurement.get("grade_gripStrength", {})
                if isinstance(grip_grade, dict):
                    grade_grip_right = grip_grade.get("rightGrip")
                    grade_grip_left = grip_grade.get("leftGrip")
                else:
                    grade_grip_right = None
                    grade_grip_left = None

                # 7. 종합 등급 계산 - 0이 아닌 등급만 포함
                all_grades = [
                    get_grade(measurement.get("50mRun"), measurement.get("grade_50mRun")),
                    get_grade(measurement.get("shuttleRun"), measurement.get("grade_shuttleRun")),
                    get_grade(measurement.get("rollingUp"), measurement.get("grade_rollingUp")),
                    get_grade(measurement.get("sitAndReach"), measurement.get("grade_sitAndReach")),
                    get_grade(measurement.get("longJump"), measurement.get("grade_longJump")),
                    get_grade(grip_right, grade_grip_right),
                    get_grade(grip_left, grade_grip_left),
                    get_grade(bmi, measurement.get("grade_bmi")),
                ]
                valid_grades = [g for g in all_grades if g and g > 0]
                grade_total = round(sum(valid_grades) / len(valid_grades)) if valid_grades else 5

                # 8. StudentData 생성
                student = StudentData(
                    name=user["name"],
                    age=user["age"],
                    gender="남" if measurement.get("gender") == "M" else "여",
                    school="",
                    class_info=f"{user.get('grade_year', '')}학년 {user.get('class_number', '')}반 {user.get('student_number', '')}번",
                    height=height,
                    weight=weight,
                    bmi=bmi,
                    run_50m=measurement.get("50mRun", 0) or 0,
                    shuttle_run=measurement.get("shuttleRun", 0) or 0,
                    sit_ups=measurement.get("rollingUp", 0) or 0,
                    flexibility=measurement.get("sitAndReach", 0) or 0,
                    long_jump=measurement.get("longJump", 0) or 0,
                    grip_right=grip_right,
                    grip_left=grip_left,
                    grade_50m=get_grade(measurement.get("50mRun"), measurement.get("grade_50mRun")),
                    grade_shuttle=get_grade(measurement.get("shuttleRun"), measurement.get("grade_shuttleRun")),
                    grade_situps=get_grade(measurement.get("rollingUp"), measurement.get("grade_rollingUp")),
                    grade_flexibility=get_grade(measurement.get("sitAndReach"), measurement.get("grade_sitAndReach")),
                    grade_long_jump=get_grade(measurement.get("longJump"), measurement.get("grade_longJump")),
                    grade_grip_right=get_grade(grip_right, grade_grip_right),
                    grade_grip_left=get_grade(grip_left, grade_grip_left),
                    grade_bmi=get_grade(bmi, measurement.get("grade_bmi")),
                    grade_total=grade_total,
                )

                # 9. PDF 생성
                os.makedirs("src/output", exist_ok=True)
                output_path = f"src/output/{user['name']}_PAPS결과지.pdf"

                generator = PapsPdfGenerator(
                    template_path="src/templates/template.pdf",
                    font_path="src/fonts"
                )
                generator.generate(student, output_path)

                # 10. ZIP에 추가
                zip_file.write(output_path, f"{user['name']}_PAPS결과지.pdf")
                pdf_count += 1

                # 임시 파일 삭제
                os.remove(output_path)

        if pdf_count == 0:
            raise HTTPException(status_code=404, detail="생성할 PDF가 없습니다. 측정 데이터를 확인해주세요.")

        zip_buffer.seek(0)
        logger.info(f"전체 PDF 생성 완료: {pdf_count}명")

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=PAPS_results_{measure_date}.zip"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"전체 PDF 생성 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))