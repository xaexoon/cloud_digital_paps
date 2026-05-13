"""엑셀 내보내기 관련 비즈니스 로직"""

from datetime import datetime
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class ExcelExportService:
    """엑셀 내보내기 서비스"""

    # 기본-1 시트 컬럼 매핑 (측정값)
    COLUMN_MAPPING_VALUE = {
        "50mRun": 4,      # D열
        "shuttleRun": 5,  # E열
        "sitAndReach": 8, # H열
        "rollingUp": 9,   # I열
        "longJump": 10,   # J열
        "height": 11,     # K열
        "weight": 12,     # L열
        "bodyMeasurement":13,
    }

    # 기본-2 시트 컬럼 매핑 (등급)
    COLUMN_MAPPING_GRADE = {
        "50mRun": 4,      # D열
        "shuttleRun": 5,  # E열
        "sitAndReach": 8, # H열
        "rollingUp": 9,   # I열
        "longJump": 10,   # J열
        "bmi": 11,        # K열
    }

    # 나이스 양식 컬럼 매핑
    NEIS_COLUMN_MAPPING = {
        "shuttleRun": 3,   # C열
        "sitAndReach": 7,  # G열
        "rollingUp": 14,   # N열
        "gripStrength": 15, # O열
        "50mRun": 16,      # P열
        "longJump": 17,    # Q열
        "bmi": 18,         # R열
    }

    def load_workbook(self, file_path: str) -> Workbook:
        """엑셀 파일 로드"""
        return load_workbook(file_path)

    def get_measure_date(self, ws) -> datetime.date:
        """기본-1 시트에서 측정일 읽기"""
        measure_cell = ws.cell(row=7, column=2)
        excel_measure_date = measure_cell.value

        if isinstance(excel_measure_date, datetime):
            return excel_measure_date.date()
        return datetime.strptime(str(excel_measure_date), "%Y-%m-%d").date()

    def build_tag_mappings(self, ws, users: List[Dict]) -> Tuple[Dict, Dict]:
        """
        태그번호 매핑 생성
        Returns: (tag_to_user, number_to_tag)
        """
        tag_to_user = {str(user["tag_number"]): user for user in users}
        number_to_tag = {}

        for row in range(10, ws.max_row + 1):
            number_cell = ws.cell(row=row, column=1)
            tag_cell = ws.cell(row=row, column=2)
            if tag_cell.value is None:
                continue

            tag = str(int(float(tag_cell.value)))
            number = str(number_cell.value).strip() if number_cell.value else None

            if number:
                number_to_tag[number] = tag

        return tag_to_user, number_to_tag

    def organize_measurement_data(self, measurements: List[Dict]) -> Dict:
        """측정 데이터를 태그번호별로 정리"""
        measurement_dict = {}
        for m in measurements:
            tag = str(m["tag_number"])
            if tag not in measurement_dict:
                measurement_dict[tag] = {"values": {}, "grades": {}}

            m_type = m["exercise_type"]

            # 악력 별도 처리
            if m_type == "gripStrength":
                measurement_dict[tag]["values"][m_type] = {
                    "leftGrip": m.get("left", 0),
                    "rightGrip": m.get("right", 0),
                }
                measurement_dict[tag]["grades"][m_type] = {
                    "leftGrip": m.get("left_grade"),
                    "rightGrip": m.get("right_grade"),
                    "grade": m.get("grade"),
                }
            # 신체측정 별도 처리
            elif m_type == "bodyMeasurement":
                measurement_dict[tag]["values"]["bodyMeasurement"] = m.get("value")  # bmi
                measurement_dict[tag]["values"]["height"] = m.get("height", 0)
                measurement_dict[tag]["values"]["weight"] = m.get("weight", 0)
                measurement_dict[tag]["grades"]["bmi"] = m.get("grade")  # 시트2 K열
            else:
                measurement_dict[tag]["values"][m_type] = m.get("value")
                measurement_dict[tag]["grades"][m_type] = m.get("grade")

        return measurement_dict

    def fill_basic1_sheet(self, ws, measurement_dict: Dict):
        """기본-1 시트에 측정값 채우기"""
        for row in range(10, ws.max_row + 1):
            tag_cell = ws.cell(row=row, column=2)
            if tag_cell.value is None:
                continue

            tag = str(int(float(tag_cell.value)))

            if tag not in measurement_dict:
                continue

            values = measurement_dict[tag]["values"]

            # gripStrength 처리 (좌/우 분리)
            self._fill_grip_values(ws, row, values)

            # 나머지 종목 처리
            for m_type, col_num in self.COLUMN_MAPPING_VALUE.items():
                if m_type in values and values[m_type] is not None:
                    ws.cell(row=row, column=col_num).value = values[m_type]

    def _fill_grip_values(self, ws, row: int, values: Dict):
        """악력 값 채우기 (시트1: 좌/우 분리)"""
        if "gripStrength" not in values:
            return

        grip_value = values["gripStrength"]
        if isinstance(grip_value, dict):
            if grip_value.get("leftGrip"):
                ws.cell(row=row, column=6).value = grip_value["leftGrip"]
            if grip_value.get("rightGrip"):
                ws.cell(row=row, column=7).value = grip_value["rightGrip"]

    def fill_basic2_sheet(self, ws_value, ws_grade, measurement_dict: Dict):
        """기본-2 시트에 등급 채우기"""
        grade_row = 3
        for value_row in range(10, ws_value.max_row + 1):
            tag_cell = ws_value.cell(row=value_row, column=2)
            if tag_cell.value is None:
                continue

            # 번호, 태그번호, 이름 복사
            ws_grade.cell(row=grade_row, column=1).value = ws_value.cell(row=value_row, column=1).value
            ws_grade.cell(row=grade_row, column=2).value = ws_value.cell(row=value_row, column=2).value
            ws_grade.cell(row=grade_row, column=3).value = ws_value.cell(row=value_row, column=3).value

            tag = str(int(float(tag_cell.value)))

            if tag in measurement_dict:
                grades = measurement_dict[tag]["grades"]
                self._fill_grip_grades(ws_grade, grade_row, grades)

                for m_type, col_num in self.COLUMN_MAPPING_GRADE.items():
                    if m_type in grades and grades[m_type] is not None:
                        ws_grade.cell(row=grade_row, column=col_num).value = grades[m_type]

            grade_row += 1

    def _fill_grip_grades(self, ws, row: int, grades: Dict):
        """악력 등급 채우기 (시트2: 좌/우 분리)"""
        if "gripStrength" not in grades:
            return

        grip_grade = grades["gripStrength"]
        if isinstance(grip_grade, dict):
            left = grip_grade.get("leftGrip")
            right = grip_grade.get("rightGrip")
            if left is not None:
                ws.cell(row=row, column=6).value = left
            if right is not None:
                ws.cell(row=row, column=7).value = right

    def fill_neis_sheet(self, ws, number_to_tag: Dict, tag_to_user: Dict, measurement_dict: Dict):
        """나이스 양식 시트 채우기"""
        for row in range(3, ws.max_row + 1):
            number_cell = ws.cell(row=row, column=1)
            if number_cell.value is None:
                continue

            number = str(number_cell.value).strip()
            tag = number_to_tag.get(number)

            if not tag:
                continue

            # 이름 복사
            if tag in tag_to_user:
                ws.cell(row=row, column=2).value = tag_to_user[tag]["name"]

            if tag not in measurement_dict:
                continue

            values = measurement_dict[tag]["values"]
            self._fill_neis_values(ws, row, values)

    def _fill_neis_values(self, ws, row: int, values: Dict):
        """나이스 양식에 값 채우기 (시트3)"""
        if "shuttleRun" in values:
            ws.cell(row=row, column=3).value = values["shuttleRun"]
        if "sitAndReach" in values:
            ws.cell(row=row, column=7).value = values["sitAndReach"]
        if "rollingUp" in values:
            ws.cell(row=row, column=14).value = values["rollingUp"]
        if "50mRun" in values:
            ws.cell(row=row, column=16).value = values["50mRun"]
        if "longJump" in values:
            ws.cell(row=row, column=17).value = values["longJump"]

        # 악력 - 좌/우 중 높은 값 (max)
        if "gripStrength" in values:
            grip_value = values["gripStrength"]
            if isinstance(grip_value, dict):
                left_grip = grip_value.get("leftGrip", 0) or 0
                right_grip = grip_value.get("rightGrip", 0) or 0
                ws.cell(row=row, column=15).value = max(left_grip, right_grip)

        # BMI 계산
        height = values.get("height")
        weight = values.get("weight")
        if height and weight and height > 0:
            height_m = height / 100
            bmi = round(weight / (height_m ** 2), 1)
            ws.cell(row=row, column=18).value = bmi


excel_export_service = ExcelExportService()