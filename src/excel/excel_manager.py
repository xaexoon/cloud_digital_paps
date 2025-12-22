# excel_manager.py

import openpyxl
from io import BytesIO


class ExcelManager:
    def __init__(self, file_contents: bytes):
        self.file_contents = file_contents
        self.wb = openpyxl.load_workbook(BytesIO(file_contents))
        self.ws = self.wb.active

    def get_school_code(self) -> str:
        """엑셀 파일에서 B2 셀의 학교코드 추출"""
        school_code = str(self.ws['B2'].value) if self.ws['B2'].value else ""
        return school_code

    def close(self):
        """워크북 닫기"""
        self.wb.close()